import json
from django.utils.http import urlencode
import openpyxl
import requests
import csv
from django.shortcuts import render
from integration_utils.bitrix24.bitrix_user_auth.main_auth import main_auth
from django.http import JsonResponse, HttpResponse
from io import StringIO, BytesIO


@main_auth(on_cookies=True)
def contacts_page(request):
    return render(request, 'contacts_page.html')


def parse_file(file):
    if file.name.endswith('.csv'):
        reader = csv.DictReader(file.read().decode('utf-8-sig').splitlines())
        return list(reader)
    elif file.name.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        return [dict(zip(headers, [cell.value for cell in row])) for row in sheet.iter_rows(min_row=2)]


def map_companies(bx, contacts):
    company_names = {c['Компания'] for c in contacts if c.get('Компания')}
    if not company_names:
        return {}

    companies = bx.call_list_method('crm.company.list', {
        'filter': {'TITLE': list(company_names)},
        'select': ['ID', 'TITLE']
    })

    map_comp = {c['TITLE']: c['ID'] for c in companies}

    missing_companies = company_names - set(map_comp.keys())

    if missing_companies:
        batch = [
            (f"create_company_{i}", "crm.company.add", {"fields": {"TITLE": name}})
            for i, name in enumerate(missing_companies)
        ]
        creation_results = bx.batch_api_call(batch)

        for name, result in zip(missing_companies, creation_results.values()):
            if result.get('result'):
                map_comp[name] = result['result']

    for c in contacts:
        c['Компания'] = map_comp[c['Компания']] if map_comp[c['Компания']] else c['Компания']

    return contacts


def process_duplicates(bx, created_contacts_ids):
    if not created_contacts_ids:
        return []
    created_contacts = bx.call_list_method('crm.contact.list',
                                           {'select': ['ID', 'EMAIL', 'PHONE'],
                                            'filter': {'@ID': created_contacts_ids}})

    batch_commands = []
    for contact in created_contacts:
        if contact.get('PHONE'):

            batch_commands.append(
                    (f"phone_contact_{contact['ID']}", 'crm.duplicate.findbycomm',
                     {'entity_type': "CONTACT",
                      'type': "PHONE",
                      'values': [contact['PHONE'][0]["VALUE"]]
                      })
            )
        if contact['EMAIL']:
            batch_commands.append(
                    (f"email_contact_{contact['ID']}", 'crm.duplicate.findbycomm',
                     {'entity_type': "CONTACT",
                      'type': "EMAIL",
                      'values': [contact['EMAIL'][0]["VALUE"]]
                      })
            )

    duplicates_results = bx.batch_api_call(batch_commands) if batch_commands else {}

    merge_results = []
    processed_groups = set()

    batch_commands = []
    other_contacts = []
    for result in duplicates_results.values():
        if not result.get('result'):
            continue
        duplicate_group = set(result['result'].get('CONTACT', []))
        if len(duplicate_group) < 2 or frozenset(duplicate_group) in processed_groups:
            continue
        processed_groups.add(frozenset(duplicate_group))

        common_contacts = duplicate_group & set(created_contacts_ids)

        main_contact_id = next(iter(common_contacts)) if common_contacts else next(iter(duplicate_group))
        other_contacts = list(duplicate_group - {main_contact_id})

        batch_commands.append(
                (f"mergeBatch{main_contact_id}-{len(other_contacts)}", 'crm.entity.mergeBatch',
                 {
                     'params': {
                         'entityTypeId': 3,
                         'entityIds': [main_contact_id] + other_contacts[:50]
                     }}))
        merge_results.append({
            'main_contact_id': main_contact_id,
            'merged_contact_ids': other_contacts
        })

    print(batch_commands)

    process_duplicates_results = bx.batch_api_call(batch_commands) if batch_commands else {}

    return merge_results


def create_contacts(bx, contacts):
    created_contacts = bx.batch_api_call([
        (f"contact_{cid}", 'crm.contact.add', {'fields': {
            'NAME': contact['Имя'],
            'LAST_NAME': contact['Фамилия'],
            'PHONE': [{'VALUE': contact['Телефон']}] if contact.get('Телефон') else None,
            'EMAIL': [{'VALUE': contact['Email']}] if contact.get('Email') else None,
            'COMPANY_ID': contact['Компания']
        }})
        for cid, contact in enumerate(contacts)
    ])

    created_contacts_ids = [result['result'] for result in created_contacts.values() if result.get('result')]
    merge_results = process_duplicates(bx, created_contacts_ids)

    return merge_results


@main_auth(on_cookies=True)
def import_file(request):
    if request.method == 'POST' and request.FILES['fileToImport']:
        bx = request.bitrix_user_token
        file = request.FILES['fileToImport']
        try:
            data = parse_file(file)
            if not data:
                return JsonResponse({"error": "Файл пуст или не содержит данных"}, status=400)
            data = map_companies(bx, data)
            merge_results = create_contacts(bx, data)

            return JsonResponse({
                "message": "Контакты добавлены",
                "merge_results": merge_results
            })
        except Exception as e:
            return JsonResponse({"error": f"Ошибка при чтении файла: {e}"}, status=400)

    return


def write_file(data, file_format='csv'):
    if file_format == 'csv':
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        response = HttpResponse(buffer.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=contacts.csv'
        return response
    elif file_format == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Контакты"

        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, row_data in enumerate(data, 2):
            for col_num, key in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data[key])

        buffer = BytesIO()
        wb.save(buffer)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=contacts.xlsx'

        return response

    return HttpResponse("Unsupported format", status=400)


@main_auth(on_cookies=True)
def export_file(request):
    if request.method == 'POST':
        try:
            bx = request.bitrix_user_token
            file_format = request.POST.get('type_id')
            date_start = request.POST.get('dateFrom')
            date_end = request.POST.get('dateTo')

            filters = {
                '!COMPANY_ID': False
            }

            if date_start:
                filters['>=DATE_CREATE'] = date_start
            if date_end:
                filters['<=DATE_CREATE'] = date_end

            contacts = bx.call_list_method('crm.contact.list', {
                'select': ['NAME', 'LAST_NAME', 'PHONE', 'EMAIL', 'COMPANY_ID'],
                'filter': filters
            })

            if contacts:
                unique_company_ids = list({c['COMPANY_ID'] for c in contacts if c['COMPANY_ID']})

                batch_commands = [
                    (f"company_{cid}", 'crm.company.get', {'id': cid})
                    for cid in unique_company_ids
                ]

                companies = bx.batch_api_call(batch_commands)

                result = []
                for contact in contacts:
                    company_data = companies["company_" + str(contact["COMPANY_ID"])]['result']
                    result.append({
                        "Имя": contact["NAME"],
                        "Фамилия": contact["LAST_NAME"],
                        "Телефон": contact["PHONE"][0]["VALUE"] if contact.get("PHONE") else None,
                        "Email": contact["EMAIL"][0]["VALUE"] if contact.get("EMAIL") else None,
                        "Компания": company_data["TITLE"]
                    })
                response = write_file(result, file_format)
                return response
            else:
                return JsonResponse({"error": "Фильтры отсеяли все контакты"}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    else:
        return JsonResponse({"error": "Неверный тип запроса"}, status=400)
